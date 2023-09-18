from django.shortcuts import render,get_object_or_404,redirect
from .forms import *
from .models import *
from django.views import generic
from django.contrib.admin.views.decorators import staff_member_required
import openpyxl
from django.http import HttpResponse


class Homepage(generic.TemplateView):
    template_name = 'homepage.html'


@staff_member_required
def staff_dashboard(request):
    department = Department.objects.all()
    student = Student.objects.all()
    department_form = DepartmentForm()

    if request.method == 'POST':
        department_form = DepartmentForm(request.POST)
        if department_form.is_valid():
            department_form.save()

    return render(request,'staff_page.html',{'department':department,'student':student,'department_form':department_form})


@staff_member_required
def student_detail(request,id):
    student = get_object_or_404(Student,id=id)
    return render(request,'student_detail.html',{'student':student})


@staff_member_required
def approve_student(request,id):
    student = get_object_or_404(Student,id=id)
    student.status = 'approved'
    student.save()
    return redirect('student_detail',id=id)


@staff_member_required
def disapprove_student(request,id):
    student = get_object_or_404(Student,id=id)
    student.status = 'disapprove'
    student.save()
    return redirect('student_detail',id=id)


def student_registration(request):
    form = StudentForm

    if request.method == 'POST':
        form = StudentForm(request.POST,request.FILES)
        if form.is_valid():
            form.save()

    return render(request,'student_form.html',{'form':form})



def generate_matriculation_number(request,id):
    # Retrieve the last generated number
    counter = MatriculationNumberCounter.objects.first()
    student = get_object_or_404(Student,id=id)
    department_code = student.department.department_code  # Replace with your department code logic
    admission_year = str(student.admission_year)[-2:]  # Get the last two digits of the admission year

    # Increment the counter and use it as a unique number
    counter.last_generated_number += 1
    counter.save()
    unique_number = str(counter.last_generated_number).zfill(3)  # Ensure it is 3 digits

    matriculation_number = f"{department_code}/{admission_year}/{unique_number}"
    student.matriculation_number = matriculation_number
    student.save()

    return redirect('student_detail',id=id)

def generate_batch_matriculation_numbers(request):
    # Retrieve the last generated number
    counter = MatriculationNumberCounter.objects.first()
    students = Student.objects.filter(matriculation_number = None)

    for student in students:
        department_code = student.department.department_code  # Replace with your department code logic
        admission_year = str(student.admission_year)[-2:]  # Get the last two digits of the admission year

        # Increment the counter and use it as a unique number
        counter.last_generated_number += 1
        counter.save()
        unique_number = str(counter.last_generated_number).zfill(3)  # Ensure it is 3 digits

        matriculation_number = f"{department_code}/{admission_year}/{unique_number}"
        student.matriculation_number = matriculation_number
        student.save()
        return redirect('staff_dashboard')



def export_student_records_as_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_record.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Student Records'

    # Write header row
    header = ['ID', 'First Name', 'Last Name' ,'Email','Department','Matriculation Number']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    queryset = Student.objects.all().values_list('id', 'first_name','last_name', 'email','department','matriculation_number')
    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num+1, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response